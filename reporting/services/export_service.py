"""
Servicio de Exportación a Excel y PDF.
Genera reportes descargables con formato profesional.
"""
import io
from datetime import datetime
from decimal import Decimal
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from weasyprint import HTML, CSS
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExportService:
    """
    Servicio para exportar reportes analíticos a Excel y PDF.
    """
    
    MONTH_NAMES = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    
    # Colores corporativos
    HEADER_COLOR = "1F2937"  # Slate 800
    HEADER_FONT_COLOR = "FFFFFF"
    ACCENT_COLOR = "3B82F6"  # Blue 500
    SUCCESS_COLOR = "10B981"  # Green 500
    WARNING_COLOR = "F59E0B"  # Amber 500
    DANGER_COLOR = "EF4444"  # Red 500
    LIGHT_BG = "F3F4F6"  # Gray 100
    
    def __init__(self, gym, report_data):
        self.gym = gym
        self.data = report_data
        self.years = report_data.get('years', [])
    
    # =========================================================================
    # EXPORTACIÓN A EXCEL
    # =========================================================================
    
    def export_to_excel(self):
        """
        Genera un archivo Excel completo con múltiples hojas.
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl no está instalado. Ejecuta: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Eliminar hoja por defecto
        wb.remove(wb.active)
        
        # Crear hojas
        self._create_summary_sheet(wb)
        self._create_revenue_sheet(wb)
        self._create_memberships_sheet(wb)
        self._create_attendance_sheet(wb)
        self._create_products_sheet(wb)
        
        # Guardar en buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    def get_excel_response(self, filename=None):
        """
        Retorna HttpResponse con el archivo Excel.
        """
        if filename is None:
            date_str = timezone.now().strftime('%Y%m%d')
            filename = f"reporte_analitico_{self.gym.slug}_{date_str}.xlsx"
        
        buffer = self.export_to_excel()
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    def _apply_header_style(self, cell):
        """Aplica estilo de cabecera a una celda."""
        cell.font = Font(bold=True, color=self.HEADER_FONT_COLOR, size=11)
        cell.fill = PatternFill(start_color=self.HEADER_COLOR, end_color=self.HEADER_COLOR, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            bottom=Side(style='thin', color='000000')
        )
    
    def _apply_number_format(self, cell, is_currency=False, is_percentage=False):
        """Aplica formato numérico a una celda."""
        if is_currency:
            cell.number_format = '#,##0.00 €'
        elif is_percentage:
            cell.number_format = '0.00%'
        else:
            cell.number_format = '#,##0'
        cell.alignment = Alignment(horizontal='right')
    
    def _apply_growth_style(self, cell, value):
        """Aplica estilo según el valor de crecimiento."""
        if value > 0:
            cell.font = Font(color=self.SUCCESS_COLOR, bold=True)
        elif value < 0:
            cell.font = Font(color=self.DANGER_COLOR, bold=True)
    
    def _auto_adjust_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_summary_sheet(self, wb):
        """Crea hoja de resumen ejecutivo."""
        ws = wb.create_sheet("Resumen Ejecutivo")
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Reporte Analítico - {self.gym.commercial_name or self.gym.name}"
        ws['A1'].font = Font(bold=True, size=16, color=self.HEADER_COLOR)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws['A2'] = f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(italic=True, color='666666')
        
        # KPIs principales
        row = 4
        ws[f'A{row}'] = "INDICADORES CLAVE (KPIs)"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Cabecera de años
        ws[f'A{row}'] = "Métrica"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        # Columna de variación
        var_col = get_column_letter(len(self.years) + 2)
        ws[f'{var_col}{row}'] = "Variación"
        self._apply_header_style(ws[f'{var_col}{row}'])
        
        row += 1
        
        # Datos de Revenue
        revenue = self.data.get('revenue', {})
        metrics = [
            ("Facturación Total", revenue.get('totals', {}), True, False),
            ("Impuestos Recaudados", {y: sum(revenue.get('taxes', {}).get(y, [])) for y in self.years}, True, False),
        ]
        
        # MRR
        mrr_data = self.data.get('mrr_arr', {})
        if mrr_data:
            metrics.append(("MRR (último mes)", {y: mrr_data.get('mrr', {}).get(y, [0])[-1] if mrr_data.get('mrr', {}).get(y) else 0 for y in self.years}, True, False))
            metrics.append(("ARR Proyectado", mrr_data.get('arr', {}), True, False))
        
        # Membresías
        membership = self.data.get('memberships', {})
        if membership:
            totals = membership.get('yearly_totals', {})
            metrics.append(("Nuevas Altas", {y: totals.get(y, {}).get('new', 0) for y in self.years}, False, False))
            metrics.append(("Cancelaciones", {y: totals.get(y, {}).get('cancelled', 0) for y in self.years}, False, False))
            metrics.append(("Churn Rate Promedio", {y: totals.get(y, {}).get('avg_churn', 0) for y in self.years}, False, True))
        
        # Asistencias
        attendance = self.data.get('attendance', {})
        if attendance:
            totals = attendance.get('yearly_totals', {})
            metrics.append(("Total Check-ins", {y: totals.get(y, {}).get('checkins', 0) for y in self.years}, False, False))
            metrics.append(("Visitantes Únicos", {y: totals.get(y, {}).get('unique', 0) for y in self.years}, False, False))
        
        # LTV
        ltv = self.data.get('ltv', {})
        if ltv:
            metrics.append(("LTV Estimado", ltv.get('ltv', {}), True, False))
        
        for metric_name, values, is_currency, is_percentage in metrics:
            ws[f'A{row}'] = metric_name
            
            prev_value = None
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                value = values.get(year, 0)
                ws[f'{col}{row}'] = value / 100 if is_percentage else value
                self._apply_number_format(ws[f'{col}{row}'], is_currency, is_percentage)
                
                if prev_value is not None and len(self.years) > 1:
                    # Calcular variación para el último año
                    if i == len(self.years) - 1 and prev_value > 0:
                        variation = ((value - prev_value) / prev_value) * 100
                        ws[f'{var_col}{row}'] = f"{variation:+.1f}%"
                        self._apply_growth_style(ws[f'{var_col}{row}'], variation)
                
                prev_value = value
            
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_revenue_sheet(self, wb):
        """Crea hoja detallada de facturación."""
        ws = wb.create_sheet("Facturación")
        
        revenue = self.data.get('revenue', {})
        
        # Título
        ws['A1'] = "FACTURACIÓN MENSUAL COMPARATIVA"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Cabecera
        ws[f'A{row}'] = "Mes"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        # Datos mensuales
        months = revenue.get('months', self.MONTH_NAMES)
        data = revenue.get('data', {})
        
        for month_idx, month in enumerate(months):
            ws[f'A{row}'] = month
            
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                value = data.get(year, [0] * 12)[month_idx]
                ws[f'{col}{row}'] = value
                self._apply_number_format(ws[f'{col}{row}'], is_currency=True)
            
            row += 1
        
        # Totales
        ws[f'A{row}'] = "TOTAL"
        ws[f'A{row}'].font = Font(bold=True)
        
        totals = revenue.get('totals', {})
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = totals.get(year, 0)
            self._apply_number_format(ws[f'{col}{row}'], is_currency=True)
            ws[f'{col}{row}'].font = Font(bold=True)
        
        # Agregar gráfico
        if len(self.years) > 0:
            chart = BarChart()
            chart.type = "col"
            chart.grouping = "clustered"
            chart.title = "Facturación por Mes"
            chart.y_axis.title = "Euros (€)"
            chart.x_axis.title = "Mes"
            
            data_ref = Reference(ws, min_col=2, min_row=3, max_col=len(self.years) + 1, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.shape = 4
            chart.width = 20
            chart.height = 10
            
            ws.add_chart(chart, f"A{row + 3}")
        
        self._auto_adjust_columns(ws)
    
    def _create_memberships_sheet(self, wb):
        """Crea hoja de membresías."""
        ws = wb.create_sheet("Membresías")
        
        membership = self.data.get('memberships', {})
        
        ws['A1'] = "MÉTRICAS DE MEMBRESÍAS"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Altas
        ws[f'A{row}'] = "NUEVAS ALTAS POR MES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Mes"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        new_data = membership.get('new_memberships', {})
        for month_idx, month in enumerate(self.MONTH_NAMES):
            ws[f'A{row}'] = month
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                ws[f'{col}{row}'] = new_data.get(year, [0] * 12)[month_idx]
            row += 1
        
        # Espacio
        row += 2
        
        # Bajas
        ws[f'A{row}'] = "CANCELACIONES POR MES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Mes"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        cancel_data = membership.get('cancellations', {})
        for month_idx, month in enumerate(self.MONTH_NAMES):
            ws[f'A{row}'] = month
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                ws[f'{col}{row}'] = cancel_data.get(year, [0] * 12)[month_idx]
            row += 1
        
        # Churn Rate
        row += 2
        ws[f'A{row}'] = "CHURN RATE MENSUAL (%)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Mes"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        churn_data = membership.get('churn_rate', {})
        for month_idx, month in enumerate(self.MONTH_NAMES):
            ws[f'A{row}'] = month
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                value = churn_data.get(year, [0] * 12)[month_idx]
                ws[f'{col}{row}'] = value / 100 if value else 0
                self._apply_number_format(ws[f'{col}{row}'], is_percentage=True)
            row += 1
        
        self._auto_adjust_columns(ws)
    
    def _create_attendance_sheet(self, wb):
        """Crea hoja de asistencias."""
        ws = wb.create_sheet("Asistencias")
        
        attendance = self.data.get('attendance', {})
        
        ws['A1'] = "MÉTRICAS DE ASISTENCIA"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Check-ins
        ws[f'A{row}'] = "CHECK-INS POR MES"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        ws[f'A{row}'] = "Mes"
        self._apply_header_style(ws[f'A{row}'])
        
        for i, year in enumerate(self.years):
            col = get_column_letter(i + 2)
            ws[f'{col}{row}'] = str(year)
            self._apply_header_style(ws[f'{col}{row}'])
        
        row += 1
        
        checkin_data = attendance.get('total_checkins', {})
        for month_idx, month in enumerate(self.MONTH_NAMES):
            ws[f'A{row}'] = month
            for i, year in enumerate(self.years):
                col = get_column_letter(i + 2)
                ws[f'{col}{row}'] = checkin_data.get(year, [0] * 12)[month_idx]
            row += 1
        
        # Gráfico de líneas para check-ins
        if len(self.years) > 0:
            chart = LineChart()
            chart.title = "Evolución de Check-ins"
            chart.y_axis.title = "Check-ins"
            chart.x_axis.title = "Mes"
            
            data_ref = Reference(ws, min_col=2, min_row=4, max_col=len(self.years) + 1, max_row=row - 1)
            cats_ref = Reference(ws, min_col=1, min_row=5, max_row=row - 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.width = 18
            chart.height = 10
            
            ws.add_chart(chart, f"A{row + 2}")
        
        self._auto_adjust_columns(ws)
    
    def _create_products_sheet(self, wb):
        """Crea hoja de productos."""
        ws = wb.create_sheet("Productos")
        
        products = self.data.get('products', {})
        
        ws['A1'] = "VENTAS DE PRODUCTOS"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Top productos por año
        for year in self.years:
            ws[f'A{row}'] = f"TOP 10 PRODUCTOS - {year}"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            ws[f'A{row}'] = "Producto"
            ws[f'B{row}'] = "Ingresos"
            ws[f'C{row}'] = "Unidades"
            self._apply_header_style(ws[f'A{row}'])
            self._apply_header_style(ws[f'B{row}'])
            self._apply_header_style(ws[f'C{row}'])
            row += 1
            
            top_products = products.get('top_products', {}).get(year, [])
            for product in top_products:
                ws[f'A{row}'] = product.get('name', 'N/A')
                ws[f'B{row}'] = product.get('revenue', 0)
                ws[f'C{row}'] = product.get('units', 0)
                self._apply_number_format(ws[f'B{row}'], is_currency=True)
                row += 1
            
            row += 2
        
        self._auto_adjust_columns(ws)
    
    # =========================================================================
    # EXPORTACIÓN A PDF
    # =========================================================================
    
    def export_to_pdf(self):
        """
        Genera un archivo PDF con el reporte.
        """
        if not PDF_AVAILABLE:
            raise ImportError("weasyprint no está instalado. Ejecuta: pip install weasyprint")
        
        # Renderizar HTML
        html_content = render_to_string('reporting/pdf_report.html', {
            'gym': self.gym,
            'data': self.data,
            'years': self.years,
            'months': self.MONTH_NAMES,
            'generated_at': timezone.now(),
        })
        
        # Convertir a PDF
        html = HTML(string=html_content)
        pdf_buffer = io.BytesIO()
        html.write_pdf(pdf_buffer)
        pdf_buffer.seek(0)
        
        return pdf_buffer
    
    def get_pdf_response(self, filename=None):
        """
        Retorna HttpResponse con el archivo PDF.
        """
        if filename is None:
            date_str = timezone.now().strftime('%Y%m%d')
            filename = f"reporte_analitico_{self.gym.slug}_{date_str}.pdf"
        
        try:
            buffer = self.export_to_pdf()
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
        except ImportError as e:
            # Si weasyprint no está disponible, usar respuesta de error
            from django.http import JsonResponse
            return JsonResponse({
                'error': 'PDF export not available',
                'message': str(e)
            }, status=501)
    
    # =========================================================================
    # EXPORTACIÓN SIMPLIFICADA (SIN DEPENDENCIAS EXTERNAS)
    # =========================================================================
    
    def export_to_csv(self, section='revenue'):
        """
        Exporta una sección específica a CSV (sin dependencias externas).
        """
        import csv
        
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        
        if section == 'revenue':
            revenue = self.data.get('revenue', {})
            
            # Cabecera
            header = ['Mes'] + [str(y) for y in self.years]
            writer.writerow(header)
            
            # Datos
            data = revenue.get('data', {})
            for month_idx, month in enumerate(self.MONTH_NAMES):
                row = [month]
                for year in self.years:
                    row.append(data.get(year, [0] * 12)[month_idx])
                writer.writerow(row)
            
            # Totales
            totals = revenue.get('totals', {})
            total_row = ['TOTAL']
            for year in self.years:
                total_row.append(totals.get(year, 0))
            writer.writerow(total_row)
        
        elif section == 'memberships':
            membership = self.data.get('memberships', {})
            
            # Cabecera
            header = ['Mes'] + [f"Altas {y}" for y in self.years] + [f"Bajas {y}" for y in self.years]
            writer.writerow(header)
            
            new_data = membership.get('new_memberships', {})
            cancel_data = membership.get('cancellations', {})
            
            for month_idx, month in enumerate(self.MONTH_NAMES):
                row = [month]
                for year in self.years:
                    row.append(new_data.get(year, [0] * 12)[month_idx])
                for year in self.years:
                    row.append(cancel_data.get(year, [0] * 12)[month_idx])
                writer.writerow(row)
        
        elif section == 'attendance':
            attendance = self.data.get('attendance', {})
            
            header = ['Mes'] + [str(y) for y in self.years]
            writer.writerow(header)
            
            checkin_data = attendance.get('total_checkins', {})
            for month_idx, month in enumerate(self.MONTH_NAMES):
                row = [month]
                for year in self.years:
                    row.append(checkin_data.get(year, [0] * 12)[month_idx])
                writer.writerow(row)
        
        buffer.seek(0)
        return buffer
    
    def get_csv_response(self, section='revenue', filename=None):
        """
        Retorna HttpResponse con el archivo CSV.
        """
        if filename is None:
            date_str = timezone.now().strftime('%Y%m%d')
            filename = f"reporte_{section}_{self.gym.slug}_{date_str}.csv"
        
        buffer = self.export_to_csv(section)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
