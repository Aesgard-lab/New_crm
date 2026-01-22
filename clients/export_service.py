"""
Servicios para exportar datos a Excel y PDF
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT


class ClientExportService:
    """
    Servicio para exportar clientes a Excel y PDF
    """
    
    @staticmethod
    def export_to_excel(clients, gym_name=""):
        """
        Exporta un queryset de clientes a Excel
        
        Args:
            clients: QuerySet de clientes
            gym_name: Nombre del gimnasio (para encabezado)
        
        Returns:
            BytesIO object con el archivo Excel
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Clientes"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título
        sheet.merge_cells('A1:H1')
        title_cell = sheet['A1']
        title_cell.value = f"Listado de Clientes - {gym_name}" if gym_name else "Listado de Clientes"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        
        # Fecha de exportación
        sheet.merge_cells('A2:H2')
        date_cell = sheet['A2']
        date_cell.value = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.font = Font(size=10, italic=True)
        date_cell.alignment = center_align
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'DNI', 'Estado', 'Fecha Creación']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Datos
        for row_num, client in enumerate(clients, 5):
            data = [
                client.id,
                client.first_name,
                client.last_name,
                client.email or '',
                client.phone_number or '',
                client.dni or '',
                client.get_status_display(),
                client.created_at.strftime('%d/%m/%Y') if client.created_at else ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Alineación especial para ID y estado
                if col_num in [1, 7]:
                    cell.alignment = center_align
        
        # Ajustar ancho de columnas
        column_widths = [8, 15, 15, 25, 15, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        
        # Guardar en BytesIO
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def export_to_pdf(clients, gym_name=""):
        """
        Exporta un queryset de clientes a PDF
        
        Args:
            clients: QuerySet de clientes
            gym_name: Nombre del gimnasio
        
        Returns:
            BytesIO object con el archivo PDF
        """
        output = io.BytesIO()
        
        # Crear documento
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        title_text = f"Listado de Clientes - {gym_name}" if gym_name else "Listado de Clientes"
        elements.append(Paragraph(title_text, title_style))
        
        # Fecha
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        elements.append(Paragraph(
            f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            date_style
        ))
        
        # Tabla
        data = [['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'DNI', 'Estado']]
        
        for client in clients:
            data.append([
                str(client.id),
                client.first_name,
                client.last_name,
                client.email or '-',
                client.phone_number or '-',
                client.dni or '-',
                client.get_status_display()
            ])
        
        # Crear tabla
        table = Table(data, colWidths=[0.6*inch, 1.2*inch, 1.2*inch, 1.8*inch, 1.2*inch, 1*inch, 1*inch])
        
        # Estilos de tabla
        table.setStyle(TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Alineación especial
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # ID
            ('ALIGN', (6, 0), (6, -1), 'CENTER'),  # Estado
        ]))
        
        elements.append(table)
        
        # Construir documento
        doc.build(elements)
        output.seek(0)
        
        return output


class ReportExportService:
    """
    Servicio para exportar reportes de membresías, pagos, etc.
    """
    
    @staticmethod
    def export_memberships_to_excel(memberships, gym_name=""):
        """Exporta membresías a Excel"""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Membresías"
        
        # Estilos
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        
        # Título
        sheet.merge_cells('A1:G1')
        title_cell = sheet['A1']
        title_cell.value = f"Reporte de Membresías - {gym_name}" if gym_name else "Reporte de Membresías"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = center_align
        
        # Encabezados
        headers = ['ID', 'Cliente', 'Plan', 'Estado', 'Fecha Inicio', 'Fecha Fin', 'Precio']
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Datos
        for row_num, membership in enumerate(memberships, 4):
            data = [
                membership.id,
                str(membership.client),
                membership.plan.name if membership.plan else '-',
                membership.get_status_display(),
                membership.start_date.strftime('%d/%m/%Y') if membership.start_date else '-',
                membership.end_date.strftime('%d/%m/%Y') if membership.end_date else '-',
                f"${membership.plan.price}" if membership.plan else '-'
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Ajustar ancho
        column_widths = [8, 20, 15, 12, 15, 15, 12]
        for i, width in enumerate(column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
