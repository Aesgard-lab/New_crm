"""
Servicio genérico de exportación a Excel y PDF
Reutilizable por todos los módulos del sistema
"""
import io
from datetime import datetime
from decimal import Decimal
from typing import List, Dict, Any, Optional, Callable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


class ExportConfig:
    """Configuración para exportación"""
    def __init__(
        self,
        title: str,
        headers: List[str],
        data_extractor: Callable[[Any], List[Any]],
        column_widths: Optional[List[int]] = None,
        brand_color: str = "#1F4788",
        landscape_mode: bool = False,
        subtitle: Optional[str] = None,
        footer_text: Optional[str] = None,
    ):
        self.title = title
        self.headers = headers
        self.data_extractor = data_extractor
        self.column_widths = column_widths or [12] * len(headers)
        self.brand_color = brand_color
        self.landscape_mode = landscape_mode
        self.subtitle = subtitle
        self.footer_text = footer_text


class GenericExportService:
    """
    Servicio genérico para exportar cualquier queryset a Excel o PDF
    """
    
    @staticmethod
    def format_value(value: Any) -> str:
        """Formatea un valor para mostrar en el reporte"""
        if value is None:
            return "-"
        if isinstance(value, bool):
            return "Sí" if value else "No"
        if isinstance(value, datetime):
            return value.strftime('%d/%m/%Y %H:%M')
        if isinstance(value, Decimal):
            return f"{value:.2f}€"
        if hasattr(value, 'strftime'):  # date
            return value.strftime('%d/%m/%Y')
        return str(value)
    
    @classmethod
    def export_to_excel(
        cls,
        queryset,
        config: ExportConfig,
        gym_name: str = ""
    ) -> io.BytesIO:
        """
        Exporta un queryset a Excel
        
        Args:
            queryset: QuerySet o lista de objetos a exportar
            config: Configuración de exportación
            gym_name: Nombre del gimnasio (opcional)
        
        Returns:
            BytesIO con el archivo Excel
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = config.title[:31]  # Excel limita a 31 caracteres
        
        # Estilos
        brand_color = config.brand_color.replace("#", "")
        header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=14, color=brand_color)
        subtitle_font = Font(size=10, italic=True, color="666666")
        border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        current_row = 1
        num_cols = len(config.headers)
        
        # Título
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        title_cell = sheet.cell(row=current_row, column=1)
        title_text = f"{config.title} - {gym_name}" if gym_name else config.title
        title_cell.value = title_text
        title_cell.font = title_font
        title_cell.alignment = center_align
        current_row += 1
        
        # Subtítulo si existe
        if config.subtitle:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
            subtitle_cell = sheet.cell(row=current_row, column=1)
            subtitle_cell.value = config.subtitle
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = center_align
            current_row += 1
        
        # Fecha de exportación
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        date_cell = sheet.cell(row=current_row, column=1)
        date_cell.value = f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.font = Font(size=9, italic=True, color="888888")
        date_cell.alignment = center_align
        current_row += 2  # Espacio extra
        
        # Encabezados
        for col_num, header in enumerate(config.headers, 1):
            cell = sheet.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        current_row += 1
        
        # Datos
        row_count = 0
        for item in queryset:
            row_data = config.data_extractor(item)
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=current_row, column=col_num)
                cell.value = cls.format_value(value)
                cell.border = border
                cell.alignment = left_align
            current_row += 1
            row_count += 1
        
        # Footer con totales si hay
        if config.footer_text:
            current_row += 1
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
            footer_cell = sheet.cell(row=current_row, column=1)
            footer_cell.value = config.footer_text
            footer_cell.font = Font(bold=True, size=10)
            footer_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Total de registros
        current_row += 1
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_cols)
        total_cell = sheet.cell(row=current_row, column=1)
        total_cell.value = f"Total de registros: {row_count}"
        total_cell.font = Font(size=9, italic=True, color="666666")
        total_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Ajustar ancho de columnas
        for i, width in enumerate(config.column_widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = width
        
        # Guardar
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return output
    
    @classmethod
    def export_to_pdf(
        cls,
        queryset,
        config: ExportConfig,
        gym_name: str = ""
    ) -> io.BytesIO:
        """
        Exporta un queryset a PDF
        
        Args:
            queryset: QuerySet o lista de objetos a exportar
            config: Configuración de exportación
            gym_name: Nombre del gimnasio (opcional)
        
        Returns:
            BytesIO con el archivo PDF
        """
        output = io.BytesIO()
        
        # Configurar página
        page_size = landscape(A4) if config.landscape_mode else A4
        
        doc = SimpleDocTemplate(
            output,
            pagesize=page_size,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.5*inch
        )
        
        elements = []
        styles = getSampleStyleSheet()
        brand_color = colors.HexColor(config.brand_color)
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=brand_color,
            spaceAfter=6,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        title_text = f"{config.title} - {gym_name}" if gym_name else config.title
        elements.append(Paragraph(title_text, title_style))
        
        # Subtítulo si existe
        if config.subtitle:
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                spaceAfter=4,
                alignment=TA_CENTER,
                fontName='Helvetica'
            )
            elements.append(Paragraph(config.subtitle, subtitle_style))
        
        # Fecha
        date_style = ParagraphStyle(
            'CustomDate',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Oblique'
        )
        elements.append(Paragraph(
            f"Exportado: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            date_style
        ))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Preparar datos de la tabla
        data = [config.headers]
        row_count = 0
        
        for item in queryset:
            row_data = config.data_extractor(item)
            data.append([cls.format_value(v) for v in row_data])
            row_count += 1
        
        # Calcular anchos de columna proporcionales
        page_width = page_size[0] - inch  # Margen
        total_width = sum(config.column_widths)
        col_widths = [(w / total_width) * page_width * 0.95 for w in config.column_widths]
        
        # Crear tabla
        table = Table(data, colWidths=col_widths, repeatRows=1)
        
        # Estilos de tabla
        table_style = TableStyle([
            # Encabezado
            ('BACKGROUND', (0, 0), (-1, 0), brand_color),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Datos
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        
        # Espacio y total
        elements.append(Spacer(1, 0.3*inch))
        
        # Footer con totales si existe
        if config.footer_text:
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=10,
                fontName='Helvetica-Bold',
                alignment=TA_RIGHT
            )
            elements.append(Paragraph(config.footer_text, footer_style))
            elements.append(Spacer(1, 0.1*inch))
        
        # Total registros
        total_style = ParagraphStyle(
            'Total',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            alignment=TA_RIGHT,
            fontName='Helvetica-Oblique'
        )
        elements.append(Paragraph(f"Total de registros: {row_count}", total_style))
        
        # Construir documento
        doc.build(elements)
        output.seek(0)
        
        return output


# ============================================================
# Configuraciones predefinidas para módulos comunes
# ============================================================

def get_staff_export_config():
    """Configuración para exportar personal"""
    return ExportConfig(
        title="Listado de Personal",
        headers=['ID', 'Nombre', 'Email', 'Rol', 'PIN', 'Estado', 'Fecha Alta'],
        data_extractor=lambda s: [
            s.id,
            f"{s.user.first_name} {s.user.last_name}",
            s.user.email,
            s.role.name if hasattr(s, 'role') and s.role else 'Sin rol',
            s.pin_code or '-',
            'Activo' if s.user.is_active else 'Inactivo',
            s.created_at if hasattr(s, 'created_at') else s.user.date_joined,
        ],
        column_widths=[8, 20, 25, 15, 10, 10, 15]
    )


def get_expense_export_config(include_totals: bool = True):
    """Configuración para exportar gastos"""
    return ExportConfig(
        title="Listado de Gastos",
        headers=['ID', 'Fecha', 'Proveedor', 'Concepto', 'Categoría', 'Base', 'IVA', 'Total', 'Estado'],
        data_extractor=lambda e: [
            e.id,
            e.issue_date,
            e.supplier.name if e.supplier else '-',
            e.concept[:40] + '...' if len(e.concept) > 40 else e.concept,
            e.category.name if e.category else '-',
            e.base_amount,
            e.tax_amount,
            e.total_amount,
            e.get_status_display(),
        ],
        column_widths=[8, 12, 18, 25, 15, 12, 10, 12, 12],
        landscape_mode=True
    )


def get_supplier_export_config():
    """Configuración para exportar proveedores"""
    return ExportConfig(
        title="Listado de Proveedores",
        headers=['ID', 'Nombre', 'CIF/NIF', 'Email', 'Teléfono', 'Dirección', 'Estado'],
        data_extractor=lambda s: [
            s.id,
            s.name,
            s.tax_id or '-',
            s.email or '-',
            s.phone or '-',
            s.address or '-',
            'Activo' if s.is_active else 'Inactivo',
        ],
        column_widths=[8, 20, 15, 25, 15, 30, 10]
    )


def get_membership_plan_export_config():
    """Configuración para exportar planes de membresía"""
    return ExportConfig(
        title="Listado de Planes de Membresía",
        headers=['ID', 'Nombre', 'Precio', 'Duración', 'Recurrente', 'Sesiones', 'Estado'],
        data_extractor=lambda p: [
            p.id,
            p.name,
            p.price,
            f"{p.duration_days} días" if p.duration_days else '-',
            'Sí' if p.is_recurring else 'No',
            p.included_sessions if hasattr(p, 'included_sessions') else '-',
            'Activo' if p.is_active else 'Inactivo',
        ],
        column_widths=[8, 25, 12, 15, 12, 12, 10]
    )


def get_product_export_config():
    """Configuración para exportar productos"""
    return ExportConfig(
        title="Listado de Productos",
        headers=['ID', 'Código', 'Nombre', 'Categoría', 'Precio', 'Stock', 'Estado'],
        data_extractor=lambda p: [
            p.id,
            p.code or '-',
            p.name,
            p.category.name if hasattr(p, 'category') and p.category else '-',
            p.price,
            p.stock if hasattr(p, 'stock') else '-',
            'Activo' if p.is_active else 'Inactivo',
        ],
        column_widths=[8, 12, 25, 18, 12, 10, 10]
    )


def get_service_export_config():
    """Configuración para exportar servicios"""
    return ExportConfig(
        title="Listado de Servicios",
        headers=['ID', 'Nombre', 'Categoría', 'Precio', 'Duración', 'Estado'],
        data_extractor=lambda s: [
            s.id,
            s.name,
            s.category.name if hasattr(s, 'category') and s.category else '-',
            s.price,
            f"{s.duration} min" if hasattr(s, 'duration') and s.duration else '-',
            'Activo' if s.is_active else 'Inactivo',
        ],
        column_widths=[8, 25, 18, 12, 12, 10]
    )


def get_activity_export_config():
    """Configuración para exportar actividades/clases"""
    return ExportConfig(
        title="Listado de Actividades",
        headers=['ID', 'Nombre', 'Categoría', 'Capacidad', 'Duración', 'Estado'],
        data_extractor=lambda a: [
            a.id,
            a.name,
            a.category.name if hasattr(a, 'category') and a.category else '-',
            a.capacity if hasattr(a, 'capacity') else '-',
            f"{a.duration} min" if hasattr(a, 'duration') and a.duration else '-',
            'Activo' if a.is_active else 'Inactivo',
        ],
        column_widths=[8, 25, 18, 12, 12, 10]
    )


def get_discount_export_config():
    """Configuración para exportar descuentos"""
    return ExportConfig(
        title="Listado de Descuentos",
        headers=['ID', 'Código', 'Nombre', 'Tipo', 'Valor', 'Válido Desde', 'Válido Hasta', 'Estado'],
        data_extractor=lambda d: [
            d.id,
            d.code,
            d.name,
            d.get_discount_type_display() if hasattr(d, 'get_discount_type_display') else '-',
            f"{d.value}%" if d.discount_type == 'PERCENTAGE' else d.value,
            d.valid_from,
            d.valid_until,
            'Activo' if d.is_active else 'Inactivo',
        ],
        column_widths=[8, 12, 20, 12, 10, 14, 14, 10]
    )
